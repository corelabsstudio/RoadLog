# -*- coding: utf-8 -*-
"""Build 2 detail images per portfolio CASE (process + Excel evidence)."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont

ROOT = Path.home() / "Desktop" / "kmong" / "portfolio"
OUT = ROOT / "images"
OUT.mkdir(parents=True, exist_ok=True)

for p in OUT.glob("*_detail*"):
    p.unlink()

FONT = next(
    p
    for p in [
        Path(r"C:\Users\hysoo\Projects\RoadLog\tools\community_poster\assets\fonts\SCDream6.otf"),
        Path(r"C:\Windows\Fonts\malgunbd.ttf"),
        Path(r"C:\Windows\Fonts\malgun.ttf"),
    ]
    if p.exists()
)


def fnt(sz: int) -> ImageFont.FreeTypeFont:
    return ImageFont.truetype(str(FONT), sz)


def text_h(d: ImageDraw.ImageDraw, text: str, font: ImageFont.FreeTypeFont) -> int:
    b = d.textbbox((0, 0), text, font=font)
    return b[3] - b[1]


def wrap(d: ImageDraw.ImageDraw, text: str, font: ImageFont.FreeTypeFont, max_w: int) -> list[str]:
    lines: list[str] = []
    cur = ""
    for ch in text:
        t = cur + ch
        if d.textbbox((0, 0), t, font=font)[2] <= max_w:
            cur = t
        else:
            if cur:
                lines.append(cur)
            cur = ch
    if cur:
        lines.append(cur)
    return lines or [""]


def draw_wrapped(
    d: ImageDraw.ImageDraw,
    x: int,
    y: int,
    text: str,
    font: ImageFont.FreeTypeFont,
    fill: tuple,
    max_w: int,
    line_gap: int = 6,
) -> int:
    for line in wrap(d, text, font, max_w):
        d.text((x, y), line, font=font, fill=fill)
        y += text_h(d, line, font) + line_gap
    return y


def section_title(d: ImageDraw.ImageDraw, x: int, y: int, text: str, accent: tuple) -> int:
    d.rectangle([x, y + 6, x + 6, y + 28], fill=accent)
    d.text((x + 18, y), text, font=fnt(26), fill=(15, 23, 42))
    return y + 46


def save_crop(img: Image.Image, y: int, path: Path, width: int = 800) -> None:
    final_h = min(max(int(y), 900), 3000)
    out = img.crop((0, 0, width, final_h))
    assert out.size[0] >= 600 and out.size[1] <= 3000
    out.save(path, "JPEG", quality=94)
    print(path.name, out.size)


def make_detail1(
    path: Path,
    case_no: str,
    title: str,
    subtitle: str,
    problem: list[str],
    pipeline: list[str],
    deliverables: list[str],
    effects: list[str],
    tech: list[str],
    accent: tuple,
) -> None:
    width, hmax = 800, 3000
    img = Image.new("RGB", (width, hmax), (248, 250, 252))
    d = ImageDraw.Draw(img)
    pad = 40
    cw = width - pad * 2
    hero = 150
    d.rectangle([0, 0, width, hero], fill=(15, 23, 42))
    d.rectangle([0, 0, 10, hero], fill=accent)
    d.text((pad, 30), f"{case_no}  ·  DETAIL 01", font=fnt(20), fill=(96, 165, 250))
    d.text((pad, 68), "문제 정의 · 처리 흐름 · 납품", font=fnt(18), fill=(148, 163, 184))
    d.text((pad, 100), "Excel Automation Portfolio", font=fnt(16), fill=(100, 116, 139))
    y = hero + 32
    y = draw_wrapped(d, pad, y, title, fnt(32), (15, 23, 42), cw, 8)
    y += 6
    y = draw_wrapped(d, pad, y, subtitle, fnt(18), (71, 85, 105), cw, 6)
    y += 22

    bh, gap = 96, 14
    half = (cw - gap) // 2
    d.rounded_rectangle([pad, y, pad + half, y + bh], radius=12, fill=(254, 242, 242))
    d.text((pad + 14, y + 14), "BEFORE", font=fnt(16), fill=(185, 28, 28))
    draw_wrapped(d, pad + 14, y + 44, "수작업 · 비표준 · 오류 위험", fnt(16), (127, 29, 29), half - 28, 4)
    d.rounded_rectangle([pad + half + gap, y, pad + cw, y + bh], radius=12, fill=(236, 253, 245))
    d.text((pad + half + gap + 14, y + 14), "AFTER", font=fnt(16), fill=(21, 128, 61))
    draw_wrapped(
        d, pad + half + gap + 14, y + 44, "규칙 기반 자동 처리", fnt(16), (20, 83, 45), half - 28, 4
    )
    y += bh + 28

    y = section_title(d, pad, y, "비즈니스 문제", accent)
    for p in problem:
        d.ellipse([pad + 4, y + 8, pad + 14, y + 18], fill=accent)
        y = draw_wrapped(d, pad + 26, y, p, fnt(18), (51, 65, 85), cw - 26, 5)
        y += 8
    y += 16

    y = section_title(d, pad, y, "처리 파이프라인", accent)
    for i, step in enumerate(pipeline, 1):
        lines = wrap(d, step, fnt(18), cw - 64)
        ch = 20 + len(lines) * 26
        d.rounded_rectangle([pad, y, pad + cw, y + ch], radius=10, fill=(255, 255, 255))
        d.rounded_rectangle([pad, y, pad + 40, y + ch], radius=10, fill=accent)
        d.rectangle([pad + 20, y, pad + 40, y + ch], fill=accent)
        d.text((pad + 8, y + ch // 2 - 10), f"{i:02d}", font=fnt(16), fill=(255, 255, 255))
        yy = y + 10
        for line in lines:
            d.text((pad + 52, yy), line, font=fnt(18), fill=(30, 41, 59))
            yy += 26
        y += ch + 10
    y += 12

    y = section_title(d, pad, y, "납품물", accent)
    for item in deliverables:
        d.rounded_rectangle([pad, y, pad + cw, y + 40], radius=8, fill=(241, 245, 249))
        d.text((pad + 14, y + 9), item, font=fnt(17), fill=(30, 41, 59))
        y += 48
    y += 10

    y = section_title(d, pad, y, "기대 효과", accent)
    for e in effects:
        d.ellipse([pad + 4, y + 8, pad + 14, y + 18], fill=(22, 163, 74))
        y = draw_wrapped(d, pad + 26, y, e, fnt(18), (51, 65, 85), cw - 26, 5)
        y += 8
    y += 14

    y = section_title(d, pad, y, "기술 키워드", accent)
    x = pad
    row_h = 0
    for t in tech:
        bb = d.textbbox((0, 0), t, font=fnt(16))
        tw, th = bb[2] - bb[0] + 24, bb[3] - bb[1] + 14
        if x + tw > pad + cw:
            x = pad
            y += row_h + 8
            row_h = 0
        d.rounded_rectangle([x, y, x + tw, y + th], radius=14, fill=(15, 23, 42))
        d.text((x + 12, y + 6), t, font=fnt(16), fill=(248, 250, 252))
        x += tw + 8
        row_h = max(row_h, th)
    y += row_h + 28
    d.rectangle([0, y, width, y + 56], fill=(15, 23, 42))
    d.text((pad, y + 16), "온해  ·  맞춤 엑셀·업무 자동화  ·  1/2", font=fnt(16), fill=(148, 163, 184))
    y += 56
    save_crop(img, y, path, width)


def sheet_preview(
    d: ImageDraw.ImageDraw,
    x: int,
    y: int,
    w: int,
    h: int,
    title: str,
    headers: list,
    rows: list,
    header_fill: tuple,
    badge: str,
    badge_fill: tuple,
) -> None:
    d.rounded_rectangle([x, y, x + w, y + h], radius=12, fill=(255, 255, 255))
    d.rectangle([x, y, x + w, y + 36], fill=(30, 41, 59))
    d.ellipse([x + 12, y + 12, x + 22, y + 22], fill=(248, 113, 113))
    d.ellipse([x + 28, y + 12, x + 38, y + 22], fill=(250, 204, 21))
    d.ellipse([x + 44, y + 12, x + 54, y + 22], fill=(74, 222, 128))
    d.text((x + 66, y + 8), title, font=fnt(14), fill=(226, 232, 240))
    bb = d.textbbox((0, 0), badge, font=fnt(13))
    bw = bb[2] - bb[0] + 16
    d.rounded_rectangle([x + w - bw - 10, y + 6, x + w - 10, y + 28], radius=8, fill=badge_fill)
    d.text((x + w - bw - 2, y + 8), badge, font=fnt(13), fill=(255, 255, 255))

    tx, ty = x + 8, y + 44
    tw = w - 16
    cols = len(headers)
    col_w = tw / max(cols, 1)
    row_h = 28
    d.rectangle([tx, ty, tx + tw, ty + row_h], fill=header_fill)
    for i, hname in enumerate(headers):
        d.text((tx + i * col_w + 6, ty + 6), str(hname)[:10], font=fnt(12), fill=(255, 255, 255))
    for ri, row in enumerate(rows):
        ry = ty + row_h + ri * row_h
        bg = (248, 250, 252) if ri % 2 == 0 else (255, 255, 255)
        d.rectangle([tx, ry, tx + tw, ry + row_h], fill=bg)
        d.line([tx, ry + row_h, tx + tw, ry + row_h], fill=(226, 232, 240), width=1)
        for ci, val in enumerate(row):
            s = str(val)
            if len(s) > 12:
                s = s[:11] + "…"
            d.text((tx + ci * col_w + 6, ry + 6), s, font=fnt(12), fill=(51, 65, 85))
    for i in range(cols + 1):
        xx = tx + i * col_w
        d.line([xx, ty, xx, ty + row_h * (1 + len(rows))], fill=(226, 232, 240), width=1)


def make_detail2(
    path: Path,
    case_no: str,
    title: str,
    note: str,
    before_block: tuple,
    after_block: tuple,
    trust_points: list[str],
    accent: tuple,
    sample_path_text: str,
) -> None:
    width, hmax = 800, 3000
    img = Image.new("RGB", (width, hmax), (241, 245, 249))
    d = ImageDraw.Draw(img)
    pad = 36
    cw = width - pad * 2
    hero = 140
    d.rectangle([0, 0, width, hero], fill=(15, 23, 42))
    d.rectangle([0, 0, 10, hero], fill=accent)
    d.text((pad, 28), f"{case_no}  ·  DETAIL 02", font=fnt(20), fill=(96, 165, 250))
    d.text((pad, 64), "실제 샘플 데이터 기반 결과 화면", font=fnt(22), fill=(248, 250, 252))
    d.text((pad, 98), "가상 데이터 · 동일 로직을 고객 파일에 적용 가능", font=fnt(15), fill=(148, 163, 184))
    y = hero + 28
    y = draw_wrapped(d, pad, y, title, fnt(28), (15, 23, 42), cw, 8)
    y += 6
    y = draw_wrapped(d, pad, y, note, fnt(16), (71, 85, 105), cw, 5)
    y += 20

    d.text((pad, y), "① 입력 (BEFORE)", font=fnt(20), fill=(185, 28, 28))
    y += 32
    st, headers, rows, hf = before_block
    table_h = 36 + 28 * (1 + len(rows)) + 16
    sheet_preview(d, pad, y, cw, table_h, st, headers, rows, hf, "BEFORE", (220, 38, 38))
    y += table_h + 24

    d.rounded_rectangle([pad + cw // 2 - 50, y, pad + cw // 2 + 50, y + 32], radius=16, fill=accent)
    d.text((pad + cw // 2 - 28, y + 6), "자동화", font=fnt(16), fill=(255, 255, 255))
    y += 48

    d.text((pad, y), "② 출력 (AFTER)", font=fnt(20), fill=(21, 128, 61))
    y += 32
    st, headers, rows, hf = after_block
    table_h = 36 + 28 * (1 + len(rows)) + 16
    sheet_preview(d, pad, y, cw, table_h, st, headers, rows, hf, "AFTER", (22, 163, 74))
    y += table_h + 28

    y = section_title(d, pad, y, "신뢰 포인트 (검증 가능)", accent)
    for t in trust_points:
        d.rounded_rectangle([pad, y, pad + cw, y + 46], radius=10, fill=(255, 255, 255))
        d.ellipse([pad + 14, y + 14, pad + 30, y + 30], fill=(22, 163, 74))
        d.text((pad + 18, y + 12), "OK", font=fnt(11), fill=(255, 255, 255))
        draw_wrapped(d, pad + 42, y + 12, t, fnt(16), (30, 41, 59), cw - 56, 4)
        y += 54
    y += 16

    d.rounded_rectangle([pad, y, pad + cw, y + 78], radius=12, fill=(255, 255, 255))
    d.text((pad + 16, y + 12), "샘플 파일 위치", font=fnt(14), fill=(100, 116, 139))
    draw_wrapped(d, pad + 16, y + 36, sample_path_text, fnt(15), (51, 65, 85), cw - 32, 4)
    y += 98

    d.rectangle([0, y, width, y + 56], fill=(15, 23, 42))
    d.text((pad, y + 16), "온해  ·  맞춤 엑셀·업무 자동화  ·  2/2", font=fnt(16), fill=(148, 163, 184))
    y += 56
    save_crop(img, y, path, width)


def read_sheet_preview(path: Path, sheet: str | None = None, max_rows: int = 5, max_cols: int = 5):
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    headers = []
    for c in range(1, max_cols + 1):
        v = ws.cell(1, c).value
        headers.append(v if v is not None else f"C{c}")
    rows = []
    for r in range(2, min(ws.max_row, 1 + max_rows) + 1):
        row = []
        for c in range(1, max_cols + 1):
            v = ws.cell(r, c).value
            if isinstance(v, float):
                v = int(v) if v == int(v) else round(v, 1)
            if isinstance(v, int) and abs(v) >= 1000:
                v = f"{v:,}"
            row.append(v if v is not None else "")
        rows.append(row)
    return headers, rows


def main() -> None:
    p1b = ROOT / "01_sheet_merge" / "BEFORE_지점별_매출_원본.xlsx"
    p1a = ROOT / "01_sheet_merge" / "AFTER_통합원장_요약.xlsx"
    p2b = ROOT / "02_daily_summary" / "BEFORE_주문RAW_3주.xlsx"
    p2a = ROOT / "02_daily_summary" / "AFTER_일별KPI_보드.xlsx"
    p3b = ROOT / "03_report_clean" / "BEFORE_경비청구_messy.xlsx"
    p3a = ROOT / "03_report_clean" / "AFTER_경비_보고용.xlsx"

    # CASE 01
    make_detail1(
        OUT / "case01_detail_1.jpg",
        "CASE 01",
        "지점별 매출 시트 자동 병합",
        "헤더·날짜 형식이 다른 시트를 통합원장과 요약 차트로 변환",
        [
            "매장마다 시트 분리, 헤더명 불일치 (일자/날짜, 매출액/매출)",
            "날짜 표기 혼재 (2026-03-01 vs 2026/03/01)",
            "마감 시 복붙으로 누락·중복·집계 오류 발생",
        ],
        [
            "시트 스캔 (메모·참고 시트 제외)",
            "헤더 매핑 테이블로 컬럼 정규화",
            "날짜 파싱 및 타입 통일",
            "통합원장 정렬 저장",
            "지점·카테고리 집계, 객단가·비중 산출",
            "지점 매출 막대 차트 생성",
        ],
        ["BEFORE_지점별_매출_원본.xlsx", "AFTER_통합원장_요약.xlsx", "run_merge.py"],
        ["마감 작업 시간 단축", "헤더 불일치로 인한 집계 누락 방지", "보고용 요약·차트 즉시 확보"],
        ["Python", "openpyxl", "헤더 매핑", "정규화", "집계", "차트"],
        (37, 99, 235),
    )
    h, r = read_sheet_preview(p1b, "강남점", 5, 5)
    ha, ra = read_sheet_preview(p1a, "지점별_요약", 4, 5)
    make_detail2(
        OUT / "case01_detail_2.jpg",
        "CASE 01",
        "샘플 엑셀 입·출력 검증 화면",
        "지점 원본 시트 → 정규화·집계 후 요약 시트로 변환된 실제 샘플",
        ("BEFORE · 강남점 시트", h, r, (22, 163, 74)),
        ("AFTER · 지점별_요약", ha, ra, (29, 78, 216)),
        [
            "헤더 불일치 시트를 동일 스키마로 병합",
            "날짜 형식 혼재 파싱 후 표준 저장",
            "지점별 총매출·주문수·객단가·비중 자동 산출",
            "재실행 스크립트(run_merge.py)로 동일 결과 재현 가능",
        ],
        (37, 99, 235),
        "Desktop\\kmong\\portfolio\\01_sheet_merge\\ (BEFORE/AFTER xlsx + 스크립트)",
    )

    # CASE 02
    make_detail1(
        OUT / "case02_detail_1.jpg",
        "CASE 02",
        "주문 RAW → 일별 KPI 보드",
        "취소 제외 규칙으로 매출·채널·상품 TOP을 한 번에 집계",
        [
            "주문 내보내기 파일은 길지만 보고 숫자는 매번 피벗으로 재작업",
            "취소 주문을 매출에 넣을지 담당자마다 기준이 다름",
            "채널·상품 성과를 따로 뽑아 시간이 많이 소요됨",
        ],
        [
            "RAW 로드 및 타입 정리",
            "상태 규칙: 취소 주문은 유효 매출에서 제외",
            "일별 주문수·유효주문·취소율·객단가 산출",
            "채널별 매출·비중 집계",
            "상품 TOP (수량·매출) 생성",
            "대시보드 한 장 요약 시트 작성",
        ],
        ["BEFORE_주문RAW_3주.xlsx", "AFTER_일별KPI_보드.xlsx", "run_daily_kpi.py"],
        ["매일·매주 동일 피벗 작업 제거", "KPI 정의를 코드로 고정해 기준 통일", "공유용 보드 즉시 생성"],
        ["Python", "openpyxl", "KPI", "비즈니스 규칙", "커머스"],
        (13, 148, 136),
    )
    h, r = read_sheet_preview(p2b, None, 5, 5)
    ha, ra = read_sheet_preview(p2a, "일별_KPI", 5, 5)
    make_detail2(
        OUT / "case02_detail_2.jpg",
        "CASE 02",
        "샘플 엑셀 입·출력 검증 화면",
        "주문 RAW → 일별 KPI 보드로 변환된 실제 샘플 (취소 제외 규칙 적용)",
        ("BEFORE · 주문RAW", h, r, (71, 85, 105)),
        ("AFTER · 일별_KPI", ha, ra, (13, 148, 136)),
        [
            "취소 주문을 유효 매출에서 제외하는 규칙 명시",
            "일별 주문수·취소율·객단가 자동 산출",
            "채널·상품 TOP 시트로 보고 범위 확장",
            "run_daily_kpi.py로 동일 집계 재현 가능",
        ],
        (13, 148, 136),
        "Desktop\\kmong\\portfolio\\02_daily_summary\\ (BEFORE/AFTER xlsx + 스크립트)",
    )

    # CASE 03
    make_detail1(
        OUT / "case03_detail_1.jpg",
        "CASE 03",
        "경비 원장 정제 · 보고 패키지",
        "금액·날짜·상태 클리닝과 중복 탐지 후 부서별 승인 합계 보고",
        [
            "금액 표기 혼재 (12,000 / 12000원 / 숫자)",
            "상태값 이명·공백 (승인, 승인완료, 대기 )",
            "날짜 형식 혼재 및 동일 청구 중복 입력으로 합계 과대",
        ],
        [
            "헤더 트림 및 컬럼 고정",
            "금액 파서 (콤마·원 접미사 제거)",
            "날짜 파서 (구분자·2자리 연도 처리)",
            "상태 표준화 (승인/대기/반려/미제출)",
            "중복 키 탐지 플래그",
            "보고 포함 규칙: 승인 AND 중복 아님",
            "부서·항목 합계 + 품질 검수 리포트",
        ],
        ["BEFORE_경비청구_messy.xlsx", "AFTER_경비_보고용.xlsx", "run_expense_clean.py"],
        ["회계 제출 전 수작업 클리닝 제거", "중복·미승인 금액 합계 혼입 방지", "검수 리포트로 숫자 근거 설명"],
        ["Python", "데이터 클리닝", "검증 규칙", "중복 탐지"],
        (217, 119, 6),
    )
    h, r = read_sheet_preview(p3b, None, 5, 5)
    ha, ra = read_sheet_preview(p3a, "정제원장", 5, 5)
    make_detail2(
        OUT / "case03_detail_2.jpg",
        "CASE 03",
        "샘플 엑셀 입·출력 검증 화면",
        "messy 경비 원장 → 정제원장·중복 플래그·보고포함 컬럼이 반영된 실제 샘플",
        ("BEFORE · 경비RAW", h, r, (120, 113, 108)),
        ("AFTER · 정제원장", ha, ra, (217, 119, 6)),
        [
            "금액·날짜·상태 비표준 값을 파서로 정규화",
            "중복 행 탐지 후 합계에서 제외",
            "승인 건만 보고 포함(Y/N)으로 표시",
            "품질검수_리포트 시트로 숫자 근거 제공",
        ],
        (217, 119, 6),
        "Desktop\\kmong\\portfolio\\03_report_clean\\ (BEFORE/AFTER xlsx + 스크립트)",
    )

    print("--- all jpg ---")
    for p in sorted(OUT.glob("*.jpg")):
        im = Image.open(p)
        print(p.name, im.size)


if __name__ == "__main__":
    main()
