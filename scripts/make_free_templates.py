# -*- coding: utf-8 -*-
"""RoadLog free resource pack: 운행일지 Excel templates + soft CTA.

Output:
  docs/marketing/free_templates/
  Desktop/로드로그_무료서식/
"""
from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "marketing" / "free_templates"
DESKTOP = Path(r"C:\Users\hysoo\Desktop\로드로그_무료서식")
SITE = "https://roadlog.co.kr"

# Brand-adjacent but print-friendly
NAVY = "0F172A"
CYAN = "0891B2"
SOFT = "ECFEFF"
LINE = "CBD5E1"
HEAD_FILL = "0E7490"
ALT_ROW = "F8FAFC"
WARN = "FEF3C7"


def thin() -> Border:
    s = Side(style="thin", color=LINE)
    return Border(left=s, right=s, top=s, bottom=s)


def style_header_row(ws, row: int, cols: int) -> None:
    fill = PatternFill("solid", fgColor=HEAD_FILL)
    font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin()


def style_input_grid(ws, start_row: int, end_row: int, cols: int) -> None:
    font = Font(name="맑은 고딕", size=10)
    for r in range(start_row, end_row + 1):
        fill = PatternFill("solid", fgColor=ALT_ROW if r % 2 == 0 else "FFFFFF")
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = font
            cell.border = thin()
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def set_widths(ws, widths: list[float]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_title_block(ws, title: str, subtitle: str, cols: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    t = ws.cell(1, 1, title)
    t.font = Font(name="맑은 고딕", bold=True, size=16, color=NAVY)
    t.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    s = ws.cell(2, 1, subtitle)
    s.font = Font(name="맑은 고딕", size=10, color="475569")
    s.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 36


def add_meta_fields(ws, row: int, labels: list[str]) -> int:
    """Label | blank | Label | blank ... on one row."""
    c = 1
    label_font = Font(name="맑은 고딕", bold=True, size=10, color=NAVY)
    for lab in labels:
        cell = ws.cell(row, c, lab)
        cell.font = label_font
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.fill = PatternFill("solid", fgColor=SOFT)
        cell.border = thin()
        blank = ws.cell(row, c + 1, "")
        blank.border = thin()
        blank.fill = PatternFill("solid", fgColor="FFFFFF")
        c += 2
    return row + 1


def add_soft_cta_banner(ws, row: int, cols: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    msg = (
        f"엑셀 작성이 번거롭다면 — 스마트폰으로 위치 스탬프 찍고 메모만 하면 "
        f"제출용으로 정리해 주는 로드로그(RoadLog) 무료 체험: {SITE}"
    )
    cell = ws.cell(row, 1, msg)
    cell.font = Font(name="맑은 고딕", size=9, color="0F766E")
    cell.fill = PatternFill("solid", fgColor="CCFBF1")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = thin()
    for c in range(1, cols + 1):
        ws.cell(row, c).border = thin()
        ws.cell(row, c).fill = PatternFill("solid", fgColor="CCFBF1")
    ws.row_dimensions[row].height = 32
    return row + 1


def add_disclaimer(ws, row: int, cols: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    text = (
        "※ 본 서식은 업무 기록·정리 편의를 위한 무료 템플릿입니다. "
        "세무·세법상 판단·공제 요건은 소속 회사 규정 및 세무 전문가 상담을 기준으로 하세요. "
        "수익·공제 보장 문구는 포함하지 않습니다."
    )
    cell = ws.cell(row, 1, text)
    cell.font = Font(name="맑은 고딕", size=8, color="64748B")
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 40
    return row + 1


def sheet_about(wb: Workbook) -> None:
    ws = wb.create_sheet("이용안내·로드로그", 0)
    set_widths(ws, [18, 72])
    ws["A1"] = "로드로그 무료 운행일지 서식 팩"
    ws["A1"].font = Font(name="맑은 고딕", bold=True, size=16, color=NAVY)
    ws.merge_cells("A1:B1")

    rows = [
        ("구성", "① 개인·업무용 일일 운행일지  ② 법인차·팀 공용 일지  ③ 월간 운행 요약"),
        ("대상", "영업·외근직, 개인차 업무 이용, 법인 차량 담당, 총무·회계 보조"),
        ("쓰는 법", "메타 정보 기입 → 일별 행 작성 → 필요 시 PDF/인쇄·메일 첨부"),
        ("소프트 안내", f"엑셀이 반복되면 로드로그에서 스탬프+메모로 자동 정리 체험: {SITE}"),
        ("체험", "카드 없이 가입 · Free 10회 · 결제 강요 없음"),
        ("톤", "광고 과잉 금지 · 필요 시에만 도구 소개 · 피드백 환영"),
        ("면책", "세무 검증·공제 보장 없음. 기록 도구/서식일 뿐."),
        ("배포", "블로그·카페·드라이브 공유 가능. 유료 재판매만 금지 권장."),
        ("문의", SITE),
    ]
    r = 3
    for k, v in rows:
        ws.cell(r, 1, k).font = Font(name="맑은 고딕", bold=True, size=11, color=HEAD_FILL)
        ws.cell(r, 1).fill = PatternFill("solid", fgColor=SOFT)
        ws.cell(r, 1).border = thin()
        ws.cell(r, 2, v).font = Font(name="맑은 고딕", size=10)
        ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="center")
        ws.cell(r, 2).border = thin()
        ws.row_dimensions[r].height = 36
        r += 1

    r += 1
    ws.cell(r, 1, "한 줄 카피 (파일·블로그 공용)")
    ws.cell(r, 1).font = Font(name="맑은 고딕", bold=True, size=11)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1
    copy = (
        "운행일지, 아직도 엑셀에 손수 적고 계신가요?\n"
        "이 서식으로 일단 깔끔히 정리하시고,\n"
        f"현장에서는 스탬프만 찍고 싶은 분들은 로드로그를 한번 써 보세요. {SITE}"
    )
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 3, end_column=2)
    cell = ws.cell(r, 1, copy)
    cell.font = Font(name="맑은 고딕", size=11, color=NAVY)
    cell.fill = PatternFill("solid", fgColor="CCFBF1")
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    for rr in range(r, r + 4):
        ws.row_dimensions[rr].height = 22


def build_daily_personal(wb: Workbook) -> None:
    ws = wb.create_sheet("01_개인업무_일일운행일지")
    cols = 9
    set_widths(ws, [12, 12, 14, 18, 18, 10, 14, 16, 18])
    add_title_block(
        ws,
        "개인·업무용 차량 운행일지 (일일)",
        "개인 차량을 업무에 쓸 때 날짜·구간·목적·거리를 남기기 위한 실무 서식입니다.",
        cols,
    )
    r = 4
    r = add_meta_fields(ws, r, ["작성자", "소속/팀", "차량번호", "기간"])
    r = add_meta_fields(ws, r, ["연료비 메모", "주차/통행 메모", "비고", "확인자"])
    r += 1
    headers = [
        "일자",
        "요일",
        "출발시각",
        "출발지",
        "도착지",
        "거리(km)",
        "업무목적",
        "동승/방문처",
        "비고",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(r, i, h)
    style_header_row(ws, r, cols)
    ws.row_dimensions[r].height = 24
    start = r + 1
    end = start + 24  # 약 1개월 분량 여유
    for i in range(start, end + 1):
        for c in range(1, cols + 1):
            ws.cell(i, c, "")
    style_input_grid(ws, start, end, cols)

    # 합계
    total_row = end + 2
    ws.cell(total_row, 5, "거리 합계(km)").font = Font(name="맑은 고딕", bold=True, size=10)
    ws.cell(total_row, 6, f"=SUM(F{start}:F{end})")
    ws.cell(total_row, 6).font = Font(name="맑은 고딕", bold=True, size=11, color=HEAD_FILL)
    ws.cell(total_row, 6).border = thin()

    r = total_row + 2
    r = add_soft_cta_banner(ws, r, cols)
    add_disclaimer(ws, r + 1, cols)

    # sample hint row colors already set
    # data validation for purpose
    dv = DataValidation(
        type="list",
        formula1='"영업방문,납품/배송,회의,교육,외근,출퇴근(업무),기타"',
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    dv.add(f"G{start}:G{end}")


def build_fleet_daily(wb: Workbook) -> None:
    ws = wb.create_sheet("02_법인차_공용일지")
    cols = 11
    set_widths(ws, [12, 14, 12, 12, 16, 16, 10, 14, 12, 12, 16])
    add_title_block(
        ws,
        "법인·공용 차량 운행일지",
        "차량 담당·총무가 여러 운전자 기록을 한 장에 모을 때 쓰는 공용 서식입니다.",
        cols,
    )
    r = 4
    r = add_meta_fields(ws, r, ["관리부서", "차량번호", "차종", "관리자"])
    r = add_meta_fields(ws, r, ["보험만기", "정기점검일", "비고", "월"])
    r += 1
    headers = [
        "일자",
        "운전자",
        "출발",
        "도착",
        "출발지",
        "도착지",
        "거리(km)",
        "업무목적",
        "시작km",
        "종료km",
        "비고/이상유무",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(r, i, h)
    style_header_row(ws, r, cols)
    start = r + 1
    end = start + 30
    for i in range(start, end + 1):
        for c in range(1, cols + 1):
            ws.cell(i, c, "")
    style_input_grid(ws, start, end, cols)
    total_row = end + 2
    ws.cell(total_row, 6, "거리 합계").font = Font(name="맑은 고딕", bold=True, size=10)
    ws.cell(total_row, 7, f"=SUM(G{start}:G{end})")
    ws.cell(total_row, 7).font = Font(name="맑은 고딕", bold=True, color=HEAD_FILL)
    ws.cell(total_row, 7).border = thin()
    r = total_row + 2
    r = add_soft_cta_banner(ws, r, cols)
    add_disclaimer(ws, r + 1, cols)


def build_monthly_summary(wb: Workbook) -> None:
    ws = wb.create_sheet("03_월간운행요약")
    cols = 8
    set_widths(ws, [14, 14, 12, 14, 18, 18, 12, 20])
    add_title_block(
        ws,
        "월간 운행·외근 요약 시트",
        "일지를 모은 뒤 월말에 결재·메일 첨부에 쓰기 좋은 한 장 요약입니다.",
        cols,
    )
    r = 4
    r = add_meta_fields(ws, r, ["성명", "부서", "차량", "해당 월"])
    r += 1
    # KPI boxes
    kpis = [
        ("총 운행거리(km)", ""),
        ("운행 일수", ""),
        ("외근/방문 건수", ""),
        ("주요 방문 Top1", ""),
    ]
    c = 1
    for lab, val in kpis:
        ws.cell(r, c, lab).font = Font(name="맑은 고딕", bold=True, size=9, color="FFFFFF")
        ws.cell(r, c).fill = PatternFill("solid", fgColor=HEAD_FILL)
        ws.cell(r, c).border = thin()
        ws.cell(r, c + 1, val).border = thin()
        ws.cell(r, c + 1).fill = PatternFill("solid", fgColor=WARN)
        c += 2
    r += 2
    headers = [
        "주차/기간",
        "운행일수",
        "거리(km)",
        "외근 건수",
        "주요 구간/방문",
        "특이사항",
        "첨부여부",
        "메모",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(r, i, h)
    style_header_row(ws, r, cols)
    start = r + 1
    end = start + 5  # 1~5주
    weeks = ["1주차", "2주차", "3주차", "4주차", "5주차", "월 합계"]
    for i, w in enumerate(weeks):
        ws.cell(start + i, 1, w)
    style_input_grid(ws, start, end, cols)
    # highlight total row
    for c in range(1, cols + 1):
        ws.cell(end, c).fill = PatternFill("solid", fgColor="E0F2FE")
        ws.cell(end, c).font = Font(name="맑은 고딕", bold=True, size=10)

    r = end + 2
    ws.cell(r, 1, "월간 한 줄 보고 (복붙용)")
    ws.cell(r, 1).font = Font(name="맑은 고딕", bold=True, size=10)
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 3, end_column=cols)
    box = ws.cell(r + 1, 1, "예) 이번 달 업무 운행 ○○km / 외근 ○건. 주요 방문: …")
    box.alignment = Alignment(wrap_text=True, vertical="top")
    box.border = thin()
    for rr in range(r + 1, r + 4):
        for c in range(1, cols + 1):
            ws.cell(rr, c).border = thin()
            ws.cell(rr, c).fill = PatternFill("solid", fgColor="FFFFFF")

    r = r + 5
    r = add_soft_cta_banner(ws, r, cols)
    add_disclaimer(ws, r + 1, cols)


def write_readme(out: Path) -> None:
    text = f"""# 로드로그 무료 운행일지 서식 팩

## 파일
- `로드로그_무료_운행일지_서식팩.xlsx` — 엑셀 3종 + 이용안내

### 시트
1. **이용안내·로드로그** — 배포 톤, 소프트 CTA, 면책
2. **01_개인업무_일일운행일지** — 개인차 업무용
3. **02_법인차_공용일지** — 팀/총무 공용
4. **03_월간운행요약** — 월말 한 장 요약

## 배포 채널 (권장 순서)
1. 네이버 블로그 (본문 + 파일 첨부/드라이브 링크)
2. 네이버 카페·지식인 답변 보조 자료
3. 구글 드라이브 / 노션 공개 링크
4. 로드로그 사이트 `/resources` 랜딩(추후)

## 카피 원칙
- ❌ 「세무사 검증」「공제 보장」「수익 보장」
- ✅ 「제출·정리하기 쉬운 실무 서식」「기록 도구」
- 소프트 CTA: 엑셀이 번거로우면 로드로그 스탬프 체험 → {SITE}

## 재생성
```bash
python scripts/make_free_templates.py
```
"""
    (out / "README.md").write_text(text, encoding="utf-8")


def write_blog_copy(out: Path) -> None:
    md = f"""# 블로그·카페 배포 원고 (무료 운행일지 서식)

톤: 실무 도움 → 자연스러운 도구 소개. 세일즈 톤 금지.

---

## 제목 후보
1. 운행일지 엑셀 서식 무료 배포 (개인업무·법인차·월간요약 3종)
2. 손글씨·엑셀 운행일지, 양식부터 정리해 두세요 (무료 다운로드)
3. 영업·외근 운행 기록 엑셀 템플릿 — 제출용으로 바로 쓰는 서식

---

## 본문 (복붙용)

안녕하세요.  
업무용으로 차를 쓰는 분들, 아직도 운행일지를 메모장·엑셀에 흩어 두고 계신가요?

월말에 몰아서 쓰다 보면 날짜가 비고,  
거리·목적·방문지가 제각각이라 정리만으로 하루가 갑니다.

그래서 **당장 인쇄·메일 첨부해도 무난한 엑셀 운행일지 서식 3종**을 묶어 두었습니다.

### 구성
1. **개인·업무용 일일 운행일지** — 개인차 업무 이용 기록  
2. **법인·공용 차량 운행일지** — 운전자·시작/종료 km 칸 포함  
3. **월간 운행 요약** — 결재·보고용 한 장  

다운로드: (드라이브 링크 또는 파일 첨부)

> 본 서식은 **업무 기록·정리 편의**를 위한 템플릿입니다.  
> 세무·세법상 판단은 회사 규정과 전문가 상담을 기준으로 해 주세요.

### 작성 팁 (짧게)
- 출발/도착을 그날그날 적기 (몰아쓰기 금지에 가까울수록 정확)
- 업무 목적은 드롭다운 수준으로 통일 (영업방문 / 회의 / 외근 등)
- 월말에 요약 시트에 주차별 합만 옮기면 보고가 빨라집니다

---

엑셀로 충분하신 분은 이 서식만으로도 정리 속도가 날 거예요.

**현장에서 매번 엑셀 켜는 게 번거롭다면**,  
스마트폰으로 위치 스탬프 찍고 짧은 메모만 남기면  
제출용 문장으로 정리해 주는 도구도 있습니다.

→ 로드로그(RoadLog) 오픈 초기 무료 체험: {SITE}  
카드 없이 가입 · Free 10회 · 먼저 써 보고 피드백 주시면 감사하겠습니다.

서식 개선 아이디어도 댓글로 남겨 주세요.  
필요한 칸이 더 있으면 반영해 업데이트하겠습니다.

---

## 짧은 캡션 (카페·오픈채)
```
운행일지 엑셀 3종 무료 나눕니다 (개인업무 / 법인공용 / 월간요약).
양식만 급하신 분 받아 가세요.
엑셀이 반복되면 스탬프로 정리하는 로드로그도 있어요 → {SITE}
```

## 지식인 답변 말미 (1~2문장)
```
기록용 엑셀 서식이 필요하시면 위 양식을 참고해 보시고,
현장 기록이 자주 끊기면 스탬프형 도구(로드로그)도 한 번 시험해 보시면 좋습니다. {SITE}
```
"""
    (out / "BLOG_COPY.md").write_text(md, encoding="utf-8")


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    DESKTOP.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # remove default
    default = wb.active
    wb.remove(default)

    sheet_about(wb)
    build_daily_personal(wb)
    build_fleet_daily(wb)
    build_monthly_summary(wb)

    xlsx_name = "로드로그_무료_운행일지_서식팩.xlsx"
    path = OUT / xlsx_name
    wb.save(path)
    shutil.copy2(path, DESKTOP / xlsx_name)

    write_readme(OUT)
    write_blog_copy(OUT)
    shutil.copy2(OUT / "README.md", DESKTOP / "README.md")
    shutil.copy2(OUT / "BLOG_COPY.md", DESKTOP / "BLOG_COPY.md")

    print("OK", path)
    print("Desktop", DESKTOP)
    print("sheets", wb.sheetnames)


if __name__ == "__main__":
    main()
