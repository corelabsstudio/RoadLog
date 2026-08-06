# -*- coding: utf-8 -*-
"""Build professional Kmong portfolio: Excel samples, scripts, case docs, cards."""
from __future__ import annotations

import random
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

ROOT = Path(r"C:\Users\hysoo\Desktop\kmong\portfolio")
DIRS = {
    "01": ROOT / "01_sheet_merge",
    "02": ROOT / "02_daily_summary",
    "03": ROOT / "03_report_clean",
    "img": ROOT / "images",
}
for p in DIRS.values():
    p.mkdir(parents=True, exist_ok=True)

FONT_PATH = next(
    p
    for p in [
        Path(r"C:\Users\hysoo\Projects\RoadLog\tools\community_poster\assets\fonts\SCDream6.otf"),
        Path(r"C:\Windows\Fonts\malgunbd.ttf"),
        Path(r"C:\Windows\Fonts\malgun.ttf"),
    ]
    if p.exists()
)

thin = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)
header_font = Font(name="Malgun Gothic", bold=True, color="FFFFFF", size=11)
title_font = Font(name="Malgun Gothic", bold=True, size=14, color="0F172A")
accent_fill = PatternFill("solid", fgColor="EFF6FF")
warn_fill = PatternFill("solid", fgColor="FEF3C7")
ok_fill = PatternFill("solid", fgColor="DCFCE7")


def fnt(size: int) -> ImageFont.FreeTypeFont:
    return ImageFont.truetype(str(FONT_PATH), size=size)


def style_header(ws, row: int, cols: int, fill_hex: str = "0F172A") -> None:
    fill = PatternFill("solid", fgColor=fill_hex)
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin


def autosize(ws, max_width: int = 28) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = max((len(str(cell.value)) for cell in col if cell.value is not None), default=8)
        ws.column_dimensions[letter].width = min(max_width, max(10, length + 2))


def money_cols(ws, min_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.number_format = "#,##0"


# ---------------------------------------------------------------------------
# CASE 01
# ---------------------------------------------------------------------------
def build_case01() -> None:
    rng = random.Random(42)
    stores = ["강남점", "홍대점", "판교점", "수원점"]
    categories = ["음료", "디저트", "MD", "세트"]
    out = DIRS["01"]
    wb = Workbook()
    wb.remove(wb.active)
    base = date(2026, 3, 1)
    all_rows: list[dict] = []

    for store in stores:
        ws = wb.create_sheet(store)
        if store == "판교점":
            headers = ["날짜", "지점", "카테고리", "매출", "건수", "담당"]
        else:
            headers = ["일자", "매장명", "품목군", "매출액", "주문수", "작성자"]
        ws.append(headers)
        style_header(ws, 1, len(headers))
        for day in range(14):
            d = base + timedelta(days=day)
            for cat in categories:
                sales = rng.randint(80_000, 1_200_000)
                cnt = rng.randint(12, 180)
                owner = rng.choice(["김민수", "이서연", "박준호", "최유진"])
                if store == "판교점":
                    ws.append([d.strftime("%Y-%m-%d"), store, cat, sales, cnt, owner])
                else:
                    ws.append([d.strftime("%Y/%m/%d"), store, cat, sales, cnt, owner])
                all_rows.append(
                    {
                        "date": d,
                        "store": store,
                        "category": cat,
                        "sales": sales,
                        "count": cnt,
                        "owner": owner,
                    }
                )
        autosize(ws)

    note = wb.create_sheet("메모_참고")
    note["A1"] = "판교점 헤더·날짜 형식 다름 / 수작업 병합 시 오류 다발"
    wb.save(out / "BEFORE_지점별_매출_원본.xlsx")

    wb2 = Workbook()
    master = wb2.active
    master.title = "통합원장"
    master.append(["날짜", "지점", "카테고리", "매출", "주문수", "담당자", "요일", "주"])
    style_header(master, 1, 8, "1D4ED8")
    all_rows.sort(key=lambda r: (r["date"], r["store"], r["category"]))
    wd = ["월", "화", "수", "목", "금", "토", "일"]
    for r in all_rows:
        master.append(
            [
                r["date"].strftime("%Y-%m-%d"),
                r["store"],
                r["category"],
                r["sales"],
                r["count"],
                r["owner"],
                wd[r["date"].weekday()],
                r["date"].isocalendar()[1],
            ]
        )
    money_cols(master, 2, 4, 5)
    autosize(master)

    sum_store = wb2.create_sheet("지점별_요약")
    sum_store.append(["지점", "총매출", "총주문수", "평균객단가", "비중(%)"])
    style_header(sum_store, 1, 5, "15803D")
    total_sales = sum(r["sales"] for r in all_rows)
    for store in stores:
        rows = [r for r in all_rows if r["store"] == store]
        s = sum(r["sales"] for r in rows)
        c = sum(r["count"] for r in rows)
        sum_store.append([store, s, c, int(s / c) if c else 0, round(s / total_sales * 100, 1)])
    money_cols(sum_store, 2, 2, 4)
    chart = BarChart()
    chart.type = "col"
    chart.title = "지점별 총매출"
    data = Reference(sum_store, min_col=2, min_row=1, max_row=1 + len(stores))
    cats = Reference(sum_store, min_col=1, min_row=2, max_row=1 + len(stores))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 15
    chart.height = 10
    sum_store.add_chart(chart, "G2")
    autosize(sum_store)

    sum_cat = wb2.create_sheet("카테고리_요약")
    sum_cat.append(["카테고리", "총매출", "총주문수"])
    style_header(sum_cat, 1, 3, "1D4ED8")
    for cat in categories:
        rows = [r for r in all_rows if r["category"] == cat]
        sum_cat.append([cat, sum(r["sales"] for r in rows), sum(r["count"] for r in rows)])
    money_cols(sum_cat, 2, 2, 3)
    autosize(sum_cat)

    guide = wb2.create_sheet("자동화_안내")
    guide["A1"] = "온해 포트폴리오 CASE 01 — 지점 시트 병합"
    guide["A1"].font = title_font
    for i, line in enumerate(
        [
            "1) 시트별 헤더 매핑으로 컬럼 정규화",
            "2) 날짜 파싱 (YYYY-MM-DD / YYYY/MM/DD)",
            "3) 통합원장 정렬 저장",
            "4) 지점·카테고리 집계 + 객단가·비중",
            "5) 지점 매출 막대 차트",
        ],
        start=3,
    ):
        guide[f"A{i}"] = line
    guide.column_dimensions["A"].width = 60
    wb2.save(out / "AFTER_통합원장_요약.xlsx")

    (out / "run_merge.py").write_text(
        '''# -*- coding: utf-8 -*-
"""CASE 01 — 지점별 시트 병합 → 통합원장 + 요약"""
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.chart import BarChart, Reference

BASE = Path(__file__).resolve().parent
SRC = BASE / "BEFORE_지점별_매출_원본.xlsx"
OUT = BASE / "AFTER_통합원장_요약_재생성.xlsx"

HEADER_MAP = {
    "날짜": "date", "일자": "date",
    "지점": "store", "매장명": "store",
    "카테고리": "category", "품목군": "category",
    "매출": "sales", "매출액": "sales",
    "건수": "count", "주문수": "count",
    "담당": "owner", "작성자": "owner",
}

def parse_date(v):
    if isinstance(v, datetime):
        return v.date()
    s = str(v).strip().replace(".", "-").replace("/", "-")
    return datetime.strptime(s[:10], "%Y-%m-%d").date()

def main():
    wb = load_workbook(SRC, data_only=True)
    rows = []
    for name in wb.sheetnames:
        if str(name).startswith("메모"):
            continue
        ws = wb[name]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        keys = [HEADER_MAP.get(str(h).strip() if h else "", None) for h in headers]
        for r in range(2, ws.max_row + 1):
            item = {}
            empty = True
            for c, key in enumerate(keys, start=1):
                if not key:
                    continue
                val = ws.cell(r, c).value
                if val not in (None, ""):
                    empty = False
                item[key] = val
            if empty:
                continue
            item["date"] = parse_date(item["date"])
            item["sales"] = int(item["sales"])
            item["count"] = int(item["count"])
            rows.append(item)

    rows.sort(key=lambda x: (x["date"], x["store"], x["category"]))
    out = Workbook()
    master = out.active
    master.title = "통합원장"
    master.append(["날짜", "지점", "카테고리", "매출", "주문수", "담당자"])
    for r in rows:
        master.append([
            r["date"].strftime("%Y-%m-%d"), r["store"], r["category"],
            r["sales"], r["count"], r["owner"],
        ])

    by_store = defaultdict(lambda: {"sales": 0, "count": 0})
    for r in rows:
        by_store[r["store"]]["sales"] += r["sales"]
        by_store[r["store"]]["count"] += r["count"]
    total = sum(v["sales"] for v in by_store.values()) or 1

    sm = out.create_sheet("지점별_요약")
    sm.append(["지점", "총매출", "총주문수", "평균객단가", "비중(%)"])
    for store, v in sorted(by_store.items()):
        avg = int(v["sales"] / v["count"]) if v["count"] else 0
        sm.append([store, v["sales"], v["count"], avg, round(v["sales"] / total * 100, 1)])

    chart = BarChart()
    chart.title = "지점별 총매출"
    data = Reference(sm, min_col=2, min_row=1, max_row=sm.max_row)
    cats = Reference(sm, min_col=1, min_row=2, max_row=sm.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    sm.add_chart(chart, "G2")
    out.save(OUT)
    print("저장 완료:", OUT, "행:", len(rows))

if __name__ == "__main__":
    main()
''',
        encoding="utf-8",
    )

    (out / "CASE.md").write_text(
        """# CASE 01 · 지점별 매출 시트 자동 병합

## 한 줄 요약
헤더·날짜 형식이 제각각인 **지점별 시트**를 정규화해 **통합원장 + 지점/카테고리 요약 + 차트**로 만듭니다.

## 비즈니스 문제
- 매장마다 시트 분리, 헤더명 불일치 (`일자`/`날짜`, `매출액`/`매출`)
- 날짜 표기 혼재 (`2026-03-01` vs `2026/03/01`)
- 마감 때 복붙 → 누락·중복·집계 오류

## 처리 파이프라인
1. 시트 스캔 (메모/참고 시트 제외)
2. 헤더 매핑 테이블로 컬럼 정규화
3. 날짜 파싱·타입 통일
4. 통합원장 정렬 저장
5. 지점·카테고리 집계, 객단가·비중
6. 막대 차트

## 납품물
- `BEFORE_지점별_매출_원본.xlsx`
- `AFTER_통합원장_요약.xlsx`
- `run_merge.py`

## 기술 키워드
Python · openpyxl · 헤더 매핑 · 데이터 정규화 · 집계 · 차트
""",
        encoding="utf-8",
    )
    print("CASE01", len(all_rows))


# ---------------------------------------------------------------------------
# CASE 02
# ---------------------------------------------------------------------------
def build_case02() -> None:
    rng2 = random.Random(7)
    channels = ["스마트스토어", "쿠팡", "자사몰", "오프라인"]
    products = [
        ("A-100", "에코텀블러"),
        ("A-210", "무선마우스"),
        ("B-020", "케이블세트"),
        ("B-330", "데스크매트"),
        ("C-110", "노트북파우치"),
    ]
    out = DIRS["02"]
    wb = Workbook()
    ws = wb.active
    ws.title = "주문RAW"
    ws.append(
        ["order_id", "ordered_at", "channel", "sku", "product_name", "qty", "unit_price", "status", "memo"]
    )
    style_header(ws, 1, 9)
    start = date(2026, 4, 1)
    raw: list[dict] = []
    oid = 10001
    for day in range(21):
        d = start + timedelta(days=day)
        for _ in range(rng2.randint(8, 18)):
            sku, pname = rng2.choice(products)
            ch = rng2.choice(channels)
            qty = rng2.randint(1, 5)
            price = rng2.choice([12900, 18900, 24900, 39000, 59000])
            status = rng2.choices(
                ["결제완료", "배송중", "구매확정", "취소"],
                weights=[20, 25, 45, 10],
            )[0]
            ts = f"{d.strftime('%Y-%m-%d')} {rng2.randint(9, 22):02d}:{rng2.randint(0, 59):02d}"
            memo = "" if rng2.random() > 0.15 else rng2.choice(["선물포장", "연락요청", "부분출고", ""])
            ws.append([oid, ts, ch, sku, pname, qty, price, status, memo])
            raw.append(
                {
                    "date": d,
                    "channel": ch,
                    "sku": sku,
                    "product": pname,
                    "qty": qty,
                    "amount": qty * price,
                    "status": status,
                }
            )
            oid += 1
    autosize(ws)
    wb.save(out / "BEFORE_주문RAW_3주.xlsx")

    valid = [r for r in raw if r["status"] != "취소"]
    cancelled = [r for r in raw if r["status"] == "취소"]
    wb2 = Workbook()
    daily = wb2.active
    daily.title = "일별_KPI"
    daily.append(["날짜", "주문수", "유효주문", "취소수", "매출(유효)", "취소송액", "취소율(%)", "객단가"])
    style_header(daily, 1, 8, "1D4ED8")
    days = sorted({r["date"] for r in raw})
    for d in days:
        day_all = [r for r in raw if r["date"] == d]
        day_ok = [r for r in day_all if r["status"] != "취소"]
        day_c = [r for r in day_all if r["status"] == "취소"]
        sales = sum(r["amount"] for r in day_ok)
        c_amt = sum(r["amount"] for r in day_c)
        n_all, n_ok, n_c = len(day_all), len(day_ok), len(day_c)
        daily.append(
            [
                d.strftime("%Y-%m-%d"),
                n_all,
                n_ok,
                n_c,
                sales,
                c_amt,
                round(n_c / n_all * 100, 1) if n_all else 0,
                int(sales / n_ok) if n_ok else 0,
            ]
        )
    money_cols(daily, 2, 5, 6)
    money_cols(daily, 2, 8, 8)
    autosize(daily)

    chs = wb2.create_sheet("채널별_요약")
    chs.append(["채널", "유효주문", "매출", "비중(%)", "평균주문금액"])
    style_header(chs, 1, 5, "15803D")
    total_sales = sum(r["amount"] for r in valid) or 1
    for ch in channels:
        rows = [r for r in valid if r["channel"] == ch]
        s = sum(r["amount"] for r in rows)
        n = len(rows)
        chs.append([ch, n, s, round(s / total_sales * 100, 1), int(s / n) if n else 0])
    money_cols(chs, 2, 3, 3)
    money_cols(chs, 2, 5, 5)
    autosize(chs)

    top = wb2.create_sheet("상품_TOP")
    top.append(["SKU", "상품명", "판매수량", "매출", "주문건수"])
    style_header(top, 1, 5, "1D4ED8")
    agg: dict = defaultdict(lambda: {"name": "", "qty": 0, "sales": 0, "orders": 0})
    for r in valid:
        a = agg[r["sku"]]
        a["name"] = r["product"]
        a["qty"] += r["qty"]
        a["sales"] += r["amount"]
        a["orders"] += 1
    for sku, a in sorted(agg.items(), key=lambda x: -x[1]["sales"]):
        top.append([sku, a["name"], a["qty"], a["sales"], a["orders"]])
    money_cols(top, 2, 4, 4)
    autosize(top)

    dash = wb2.create_sheet("대시보드_요약")
    dash["A1"] = "주문 데이터 일별 KPI 보드 (자동화 결과)"
    dash["A1"].font = title_font
    valid_sales = sum(r["amount"] for r in valid)
    metrics = [
        ("기간", f"{days[0]} ~ {days[-1]}"),
        ("총 주문(RAW)", len(raw)),
        ("유효 주문", len(valid)),
        ("취소 주문", len(cancelled)),
        ("유효 매출 합계", f"{valid_sales:,}원"),
        ("전체 취소율", f"{round(len(cancelled) / len(raw) * 100, 1)}%"),
        ("일 평균 유효매출", f"{int(valid_sales / len(days)):,}원"),
    ]
    dash["A3"] = "지표"
    dash["B3"] = "값"
    style_header(dash, 3, 2)
    for i, (k, v) in enumerate(metrics, start=4):
        dash[f"A{i}"] = k
        dash[f"B{i}"] = v
        if i % 2 == 0:
            dash[f"A{i}"].fill = accent_fill
            dash[f"B{i}"].fill = accent_fill
    dash.column_dimensions["A"].width = 22
    dash.column_dimensions["B"].width = 36

    guide = wb2.create_sheet("자동화_안내")
    guide["A1"] = "온해 포트폴리오 CASE 02 — 주문 RAW KPI"
    guide["A1"].font = title_font
    for i, line in enumerate(
        ["취소 제외 유효매출 규칙", "일별 KPI", "채널 비중", "상품 TOP", "대시보드 요약"],
        start=3,
    ):
        guide[f"A{i}"] = line
    wb2.save(out / "AFTER_일별KPI_보드.xlsx")

    (out / "run_daily_kpi.py").write_text(
        '''# -*- coding: utf-8 -*-
"""CASE 02 — 주문 RAW → 일별 KPI 보드"""
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from openpyxl import load_workbook, Workbook

BASE = Path(__file__).resolve().parent
SRC = BASE / "BEFORE_주문RAW_3주.xlsx"
OUT = BASE / "AFTER_일별KPI_보드_재생성.xlsx"

def main():
    wb = load_workbook(SRC, data_only=True)
    ws = wb.active
    rows = []
    for r in range(2, ws.max_row + 1):
        status = ws.cell(r, 8).value
        qty = int(ws.cell(r, 6).value)
        price = int(ws.cell(r, 7).value)
        ts = str(ws.cell(r, 2).value)
        d = datetime.strptime(ts[:10], "%Y-%m-%d").date()
        rows.append({
            "date": d,
            "channel": ws.cell(r, 3).value,
            "sku": ws.cell(r, 4).value,
            "product": ws.cell(r, 5).value,
            "qty": qty,
            "amount": qty * price,
            "status": status,
        })

    out = Workbook()
    daily = out.active
    daily.title = "일별_KPI"
    daily.append(["날짜", "주문수", "유효주문", "취소수", "매출(유효)", "취소율(%)", "객단가"])
    for d in sorted({r["date"] for r in rows}):
        all_ = [x for x in rows if x["date"] == d]
        ok = [x for x in all_ if x["status"] != "취소"]
        c = len(all_) - len(ok)
        sales = sum(x["amount"] for x in ok)
        aov = int(sales / len(ok)) if ok else 0
        daily.append([
            d.isoformat(), len(all_), len(ok), c, sales,
            round(c / len(all_) * 100, 1) if all_ else 0, aov,
        ])

    chs = out.create_sheet("채널별_요약")
    chs.append(["채널", "유효주문", "매출"])
    by = defaultdict(lambda: {"n": 0, "s": 0})
    for x in rows:
        if x["status"] == "취소":
            continue
        by[x["channel"]]["n"] += 1
        by[x["channel"]]["s"] += x["amount"]
    for ch, v in sorted(by.items(), key=lambda i: -i[1]["s"]):
        chs.append([ch, v["n"], v["s"]])

    out.save(OUT)
    print("저장 완료:", OUT)

if __name__ == "__main__":
    main()
''',
        encoding="utf-8",
    )

    (out / "CASE.md").write_text(
        """# CASE 02 · 주문 RAW → 일별 KPI 보드

## 한 줄 요약
커머스/매장 **주문 원장(RAW)** 을 규칙 기반으로 집계해 **일별 KPI · 채널 · 상품 TOP** 보고 보드로 변환합니다.

## 비즈니스 문제
- 내보내기 파일은 길지만, 보고 숫자는 매번 피벗으로 다시 만듦
- 취소 주문을 매출에 넣을지 담당자마다 기준이 다름
- 채널·상품 성과를 따로 뽑아 시간이  lag

## 처리 파이프라인
1. RAW 로드 및 타입 정리
2. 상태 규칙: `취소` 제외한 유효 매출 정의
3. 일별 주문수·유효주문·취소율·객단가
4. 채널별 매출·비중
5. 상품 TOP (수량·매출)
6. 대시보드 한 장 요약

## 납품물
- `BEFORE_주문RAW_3주.xlsx`
- `AFTER_일별KPI_보드.xlsx`
- `run_daily_kpi.py`

## 기술 키워드
Python · openpyxl · KPI · 비즈니스 규칙 · 커머스 데이터
""",
        encoding="utf-8",
    )
    print("CASE02", len(raw))


# ---------------------------------------------------------------------------
# CASE 03
# ---------------------------------------------------------------------------
def build_case03() -> None:
    rng3 = random.Random(99)
    depts = ["영업", "마케팅", "개발", "운영"]
    types = ["교통비", "식대", "소모품", "소프트웨어", "출장"]
    people = ["강하늘", "문지호", "윤서아", "한도윤", "배예린", "오세준"]
    out = DIRS["03"]
    wb = Workbook()
    ws = wb.active
    ws.title = "경비RAW"
    ws.append(["No", " 날짜 ", "부서", "이름", "항목", "금액", "통화", "상태", "비고", "승인"])
    style_header(ws, 1, 10)
    start = date(2026, 5, 1)
    rows: list[dict] = []
    no = 1
    for _ in range(60):
        d = start + timedelta(days=rng3.randint(0, 25))
        dept = rng3.choice(depts)
        name = rng3.choice(people)
        t = rng3.choice(types)
        amount = rng3.choice([3500, 8000, 12000, 15600, 22000, 45000, 89000, 120000])
        if rng3.random() < 0.12:
            amount_cell: int | str = f"{amount:,}"
        elif rng3.random() < 0.08:
            amount_cell = f"{amount}원"
        else:
            amount_cell = amount
        status = rng3.choices(
            ["대기", "승인", "반려", "대기 ", "승인완료", "미제출"],
            weights=[25, 40, 10, 5, 15, 5],
        )[0]
        cur = rng3.choices(["KRW", "원", "KRW ", "krw"], weights=[50, 20, 15, 15])[0]
        note = rng3.choice(["", "택시", "팀회식", "AWS", "인쇄", "거래처미팅", "  "])
        appr = "" if ("대기" in status or status == "미제출") else rng3.choice(["김팀장", "박실장", ""])
        if rng3.random() < 0.2:
            d_cell = d.strftime("%y.%m.%d")
        elif rng3.random() < 0.3:
            d_cell = d.strftime("%Y/%m/%d")
        else:
            d_cell = d.strftime("%Y-%m-%d")
        ws.append([no, d_cell, dept, name, t, amount_cell, cur, status, note, appr])
        rows.append(
            {
                "date": d,
                "dept": dept,
                "name": name,
                "type": t,
                "amount": amount,
                "status_raw": status,
                "note": note.strip(),
            }
        )
        no += 1

    dup = rows[10]
    ws.append(
        [
            no,
            dup["date"].strftime("%Y-%m-%d"),
            dup["dept"],
            dup["name"],
            dup["type"],
            dup["amount"],
            "KRW",
            "승인",
            "중복의심",
            "김팀장",
        ]
    )
    rows.append({**dup, "status_raw": "승인", "note": "중복의심"})
    autosize(ws)
    wb.save(out / "BEFORE_경비청구_messy.xlsx")

    def norm_status(s: str) -> str:
        s = str(s).strip()
        if s in ("승인", "승인완료"):
            return "승인"
        if "반려" in s:
            return "반려"
        if "미제출" in s:
            return "미제출"
        return "대기"

    clean: list[dict] = []
    seen: set = set()
    duplicates = 0
    for r in rows:
        st = norm_status(r["status_raw"])
        key = (r["date"], r["dept"], r["name"], r["type"], r["amount"], st)
        is_dup = key in seen
        if is_dup:
            duplicates += 1
        else:
            seen.add(key)
        clean.append({**r, "status": st, "is_dup": is_dup})

    wb2 = Workbook()
    cws = wb2.active
    cws.title = "정제원장"
    cws.append(["날짜", "부서", "이름", "항목", "금액", "상태", "비고", "중복플래그", "보고포함"])
    style_header(cws, 1, 9, "1D4ED8")
    for r in sorted(clean, key=lambda x: (x["date"], x["dept"], x["name"])):
        include = "Y" if (r["status"] == "승인" and not r["is_dup"]) else "N"
        flag = "DUPLICATE" if r["is_dup"] else ""
        cws.append(
            [
                r["date"].strftime("%Y-%m-%d"),
                r["dept"],
                r["name"],
                r["type"],
                r["amount"],
                r["status"],
                r["note"],
                flag,
                include,
            ]
        )
        if flag:
            for col in range(1, 10):
                cws.cell(cws.max_row, col).fill = warn_fill
        elif include == "Y":
            cws.cell(cws.max_row, 9).fill = ok_fill
    money_cols(cws, 2, 5, 5)
    autosize(cws)

    report = wb2.create_sheet("부서별_승인합계")
    report.append(["부서", "승인건수", "승인금액", "대기건수", "반려건수"])
    style_header(report, 1, 5, "15803D")
    for dept in depts:
        rs = [r for r in clean if r["dept"] == dept and not r["is_dup"]]
        appr = [r for r in rs if r["status"] == "승인"]
        wait = [r for r in rs if r["status"] == "대기"]
        rej = [r for r in rs if r["status"] == "반려"]
        report.append([dept, len(appr), sum(r["amount"] for r in appr), len(wait), len(rej)])
    money_cols(report, 2, 3, 3)
    autosize(report)

    by_type = wb2.create_sheet("항목별_승인합계")
    by_type.append(["항목", "승인건수", "승인금액"])
    style_header(by_type, 1, 3, "1D4ED8")
    for t in types:
        rs = [r for r in clean if r["type"] == t and r["status"] == "승인" and not r["is_dup"]]
        by_type.append([t, len(rs), sum(r["amount"] for r in rs)])
    money_cols(by_type, 2, 3, 3)
    autosize(by_type)

    qa = wb2.create_sheet("품질검수_리포트")
    qa["A1"] = "데이터 품질 · 회계 제출 전 검수"
    qa["A1"].font = title_font
    approved_total = sum(r["amount"] for r in clean if r["status"] == "승인" and not r["is_dup"])
    checks = [
        ("원본 행 수", len(rows)),
        ("중복 의심 행", duplicates),
        ("금액 문자열 혼재 처리", "콤마·원 접미사 정규화 완료"),
        ("상태값 표준화", "승인/대기/반려/미제출"),
        ("보고 포함 기준", "승인 + 중복 아님"),
        ("승인 총액(보고)", f"{approved_total:,}원"),
    ]
    qa["A3"] = "검수항목"
    qa["B3"] = "결과"
    style_header(qa, 3, 2)
    for i, (k, v) in enumerate(checks, start=4):
        qa[f"A{i}"] = k
        qa[f"B{i}"] = v
    qa.column_dimensions["A"].width = 24
    qa.column_dimensions["B"].width = 40

    guide = wb2.create_sheet("자동화_안내")
    guide["A1"] = "온해 포트폴리오 CASE 03 — 경비 정제/보고"
    guide["A1"].font = title_font
    for i, line in enumerate(
        ["금액·날짜 파서", "상태 표준화", "중복 탐지", "부서/항목 합계", "품질 검수 리포트"],
        start=3,
    ):
        guide[f"A{i}"] = line
    wb2.save(out / "AFTER_경비_보고용.xlsx")

    (out / "run_expense_clean.py").write_text(
        '''# -*- coding: utf-8 -*-
"""CASE 03 — 경비 messy → 보고용 정제"""
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from openpyxl import load_workbook, Workbook

BASE = Path(__file__).resolve().parent
SRC = BASE / "BEFORE_경비청구_messy.xlsx"
OUT = BASE / "AFTER_경비_보고용_재생성.xlsx"

def parse_amount(v):
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).replace(",", "").replace("원", "").replace(" ", "")
    return int(float(s))

def parse_date(v):
    if isinstance(v, datetime):
        return v.date()
    s = str(v).strip().replace(".", "-").replace("/", "-")
    parts = s.split("-")
    if len(parts[0]) == 2:
        s = "20" + s
    return datetime.strptime(s[:10], "%Y-%m-%d").date()

def norm_status(s):
    s = str(s).strip()
    if s in ("승인", "승인완료"):
        return "승인"
    if "반려" in s:
        return "반려"
    if "미제출" in s:
        return "미제출"
    return "대기"

def main():
    wb = load_workbook(SRC, data_only=True)
    ws = wb.active
    rows = []
    for r in range(2, ws.max_row + 1):
        rows.append({
            "date": parse_date(ws.cell(r, 2).value),
            "dept": str(ws.cell(r, 3).value).strip(),
            "name": str(ws.cell(r, 4).value).strip(),
            "type": str(ws.cell(r, 5).value).strip(),
            "amount": parse_amount(ws.cell(r, 6).value),
            "status": norm_status(ws.cell(r, 8).value),
            "note": str(ws.cell(r, 9).value or "").strip(),
        })
    seen = set()
    for r in rows:
        key = (r["date"], r["dept"], r["name"], r["type"], r["amount"], r["status"])
        r["dup"] = key in seen
        if not r["dup"]:
            seen.add(key)

    out = Workbook()
    c = out.active
    c.title = "정제원장"
    c.append(["날짜", "부서", "이름", "항목", "금액", "상태", "중복", "보고포함"])
    for r in rows:
        inc = "Y" if r["status"] == "승인" and not r["dup"] else "N"
        c.append([
            r["date"].isoformat(), r["dept"], r["name"], r["type"], r["amount"],
            r["status"], "Y" if r["dup"] else "", inc,
        ])

    dep = out.create_sheet("부서별_승인합계")
    dep.append(["부서", "승인금액"])
    agg = defaultdict(int)
    for r in rows:
        if r["status"] == "승인" and not r["dup"]:
            agg[r["dept"]] += r["amount"]
    for k, v in sorted(agg.items()):
        dep.append([k, v])

    out.save(OUT)
    print("저장 완료:", OUT)

if __name__ == "__main__":
    main()
''',
        encoding="utf-8",
    )

    (out / "CASE.md").write_text(
        """# CASE 03 · 경비 원장 정제 → 회계·보고 패키지

## 한 줄 요약
금액·날짜·상태 표기가 지저분한 **경비 청구 RAW** 를 표준화하고, 중복을 걸러 **부서/항목 승인 합계** 보고서로 만듭니다.

## 비즈니스 문제
- `12,000` / `12000원` / 숫자 혼재
- 상태값 `승인`·`승인완료`·`대기 ` 공백·이명 표기
- 날짜 `25.05.01` / `2025/05/01` 혼재
- 같은 청구 중복 입력 → 합계 과대

## 처리 파이프라인
1. 헤더 트림·컬럼 고정
2. 금액 파서 (콤마·원 제거)
3. 날짜 파서 (구분자·2자리 연도)
4. 상태 표준화 (승인/대기/반려/미제출)
5. 중복 키 탐지 플래그
6. 보고 포함 규칙: 승인 ∧ ¬중복
7. 부서별·항목별 합계 + 품질 검수 시트

## 납품물
- `BEFORE_경비청구_messy.xlsx`
- `AFTER_경비_보고용.xlsx`
- `run_expense_clean.py`

## 기술 키워드
Python · 데이터 클리닝 · 검증 규칙 · 중복 탐지 · 관리회계 리포트
""",
        encoding="utf-8",
    )
    print("CASE03", len(rows), "dups", duplicates)


# ---------------------------------------------------------------------------
# Images + index
# ---------------------------------------------------------------------------
def draw_card(
    path: Path,
    case_no: str,
    title: str,
    subtitle: str,
    bullets: list[str],
    tags: list[str],
    accent: tuple[int, int, int],
) -> None:
    w, h = 1200, 900
    img = Image.new("RGB", (w, h), (248, 250, 252))
    d = ImageDraw.Draw(img)
    d.rectangle([0, 0, 16, h], fill=accent)
    d.rectangle([0, 0, w, 96], fill=(15, 23, 42))
    d.text((40, 28), "ONHAE  ·  Excel Automation Portfolio", font=fnt(28), fill=(226, 232, 240))
    d.text((w - 200, 34), case_no, font=fnt(26), fill=(96, 165, 250))
    d.text((40, 130), title, font=fnt(44), fill=(15, 23, 42))
    d.text((40, 200), subtitle, font=fnt(24), fill=(71, 85, 105))

    box_y, box_h, gap = 280, 200, 30
    box_w = (w - 80 - gap) // 2
    d.rounded_rectangle([40, box_y, 40 + box_w, box_y + box_h], radius=16, fill=(254, 242, 242))
    d.text((60, box_y + 20), "BEFORE", font=fnt(22), fill=(185, 28, 28))
    d.text((60, box_y + 70), "수작업 · 비표준 · 오류 위험", font=fnt(26), fill=(127, 29, 29))
    d.text((60, box_y + 120), "반복 복붙 / 피벗 / 눈집계", font=fnt(22), fill=(153, 27, 27))
    x2 = 40 + box_w + gap
    d.rounded_rectangle([x2, box_y, x2 + box_w, box_y + box_h], radius=16, fill=(236, 253, 245))
    d.text((x2 + 20, box_y + 20), "AFTER", font=fnt(22), fill=(21, 128, 61))
    d.text((x2 + 20, box_y + 70), "규칙 기반 자동 처리", font=fnt(26), fill=(20, 83, 45))
    d.text((x2 + 20, box_y + 120), "원장 + 요약 + 검증 리포트", font=fnt(22), fill=(22, 101, 52))

    y = 520
    d.text((40, y), "처리 포인트", font=fnt(26), fill=(15, 23, 42))
    y += 50
    for b in bullets:
        d.ellipse([48, y + 10, 60, y + 22], fill=accent)
        d.text((76, y), b, font=fnt(24), fill=(51, 65, 85))
        y += 42

    y = h - 90
    x = 40
    for t in tags:
        bbox = d.textbbox((0, 0), t, font=fnt(20))
        tw = bbox[2] - bbox[0] + 28
        d.rounded_rectangle([x, y, x + tw, y + 40], radius=20, fill=(15, 23, 42))
        d.text((x + 14, y + 8), t, font=fnt(20), fill=(248, 250, 252))
        x += tw + 12

    img.save(path)
    img.convert("RGB").save(path.with_suffix(".jpg"), quality=95)
    print("card", path.name)


def build_images() -> None:
    img_dir = DIRS["img"]
    draw_card(
        img_dir / "case01_card.png",
        "CASE 01",
        "지점별 매출 시트 자동 병합",
        "헤더·날짜 형식이 다른 시트를 통합원장과 요약 차트로",
        [
            "시트별 헤더 매핑으로 컬럼 정규화",
            "날짜 파서 (하이픈/슬래시 혼재 처리)",
            "지점·카테고리 집계 + 객단가·비중",
            "지점 매출 막대 차트 자동 생성",
        ],
        ["병합", "정규화", "집계", "차트"],
        (37, 99, 235),
    )
    draw_card(
        img_dir / "case02_card.png",
        "CASE 02",
        "주문 RAW → 일별 KPI 보드",
        "취소 제외 규칙으로 매출·채널·상품 TOP을 한 번에",
        [
            "상태 규칙: 취소 주문 매출 제외",
            "일별 주문수·취소율·객단가",
            "채널별 매출 비중 산출",
            "상품 TOP 및 대시보드 요약 시트",
        ],
        ["KPI", "커머스", "채널", "대시보드"],
        (13, 148, 136),
    )
    draw_card(
        img_dir / "case03_card.png",
        "CASE 03",
        "경비 원장 정제 · 보고 패키지",
        "금액·날짜·상태 클리닝 + 중복 탐지 후 부서 합계",
        [
            "금액 파서 (콤마·원 접미사 제거)",
            "상태 표준화 (승인/대기/반려)",
            "중복 키 탐지 및 합계 제외",
            "부서·항목 승인합계 + 품질 검수 리포트",
        ],
        ["클리닝", "검증", "중복", "관리회계"],
        (217, 119, 6),
    )

    w, h = 1200, 675
    img = Image.new("RGB", (w, h), (15, 23, 42))
    d = ImageDraw.Draw(img)
    d.text((50, 60), "엑셀 · 업무 자동화 포트폴리오", font=fnt(42), fill=(248, 250, 252))
    d.text((50, 130), "온해  |  Python 기반 맞춤 자동화", font=fnt(26), fill=(148, 163, 184))
    items = [
        ("01", "시트 병합", "지점 데이터 통합"),
        ("02", "KPI 보드", "주문·매출 지표"),
        ("03", "데이터 정제", "경비·보고 패키지"),
    ]
    x = 50
    for no, t, s in items:
        d.rounded_rectangle([x, 240, x + 340, 480], radius=20, fill=(30, 41, 59))
        d.text((x + 28, 270), no, font=fnt(28), fill=(96, 165, 250))
        d.text((x + 28, 330), t, font=fnt(34), fill=(248, 250, 252))
        d.text((x + 28, 400), s, font=fnt(22), fill=(148, 163, 184))
        x += 370
    d.text((50, 560), "Before/After 샘플 · 재실행 스크립트 · 규칙 문서화", font=fnt(22), fill=(100, 116, 139))
    img.save(img_dir / "portfolio_overview.png")
    img.save(img_dir / "portfolio_overview.jpg", quality=95)
    print("overview saved")


def write_index() -> None:
    (ROOT / "README.md").write_text(
        """# 온해 · 엑셀 업무 자동화 포트폴리오

전문 지식이 드러나도록 **실제 업무에서 반복되는 3가지 패턴**으로 구성했습니다.  
데이터는 모두 **가상(픽션)** 이며, 동일 로직을 고객 파일에 맞춰 재적용할 수 있습니다.

## 구성

| CASE | 주제 | 핵심 역량 |
|------|------|-----------|
| 01 | 지점별 시트 **자동 병합** | 헤더 매핑, 정규화, 집계, 차트 |
| 02 | 주문 RAW → **일별 KPI** | 비즈니스 규칙, 커머스 지표 |
| 03 | 경비 **정제·검증·보고** | 파싱, 중복 탐지, 품질 리포트 |

## 폴더

```
portfolio/
  images/                 크몽 업로드용 카드·오버뷰
  01_sheet_merge/         BEFORE / AFTER / run_merge.py / CASE.md
  02_daily_summary/
  03_report_clean/
  README.md
  KMONG_PORTFOLIO_TEXT.md
```

## 크몽 업로드 순서

1. `images/portfolio_overview.jpg`
2. `images/case01_card.jpg` (+ AFTER 엑셀 화면 캡처 권장)
3. `images/case02_card.jpg`
4. `images/case03_card.jpg`

## 로컬 검증

```powershell
cd $env:USERPROFILE\\Desktop\\kmong\\portfolio\\01_sheet_merge
python run_merge.py
```

openpyxl 필요: `pip install openpyxl`
""",
        encoding="utf-8",
    )

    (ROOT / "KMONG_PORTFOLIO_TEXT.md").write_text(
        """# 크몽 포트폴리오·서비스 소개 복붙 문구

## 포트폴리오 제목
1. 지점별 매출 시트 헤더 정규화 후 통합원장·요약 차트 자동화
2. 주문 RAW 기반 일별 KPI·채널·상품 TOP 보드 구축
3. 경비 데이터 클리닝·중복 탐지·부서별 승인 합계 보고서

## 서비스 상세 본문

반복되는 엑셀 작업, 규칙만 정리되면 자동화할 수 있습니다.

파이썬(openpyxl)으로 아래와 같은 **맞춤 자동화**를 제작합니다.

✔ 여러 시트·파일 병합 (헤더·날짜 형식 불일치 포함)
✔ 일별/채널별 KPI 집계 (취소 제외 등 비즈니스 규칙 반영)
✔ 지저분한 원장 정제 (금액·상태 표준화, 중복 탐지)
✔ 보고용 요약 시트·간단 차트
✔ 재실행 가능한 스크립트 + 사용 방법 안내

### 진행 방식
1. 샘플 파일(또는 마스킹 데이터)과 원하는 결과 형태 확인
2. 처리 규칙 문서화 (포함/제외 기준)
3. 스크립트 제작 및 결과 파일 검증
4. 납품: 결과 xlsx + 실행 방법 + 기본 수정

### 문의 시 알려주세요
1) 현재 파일 형태(시트 수, 대략 행 수)
2) 자동화하고 싶은 결과
3) 희망 기한
""",
        encoding="utf-8",
    )


def main() -> None:
    build_case01()
    build_case02()
    build_case03()
    build_images()
    write_index()
    print("ALL DONE ->", ROOT)


if __name__ == "__main__":
    main()
